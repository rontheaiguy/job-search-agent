"""
Rescore existing Job_Tracker.xlsx rows under the anchored rubric (M2).

Usage:
  python3 rescore.py --sample     # score ~12 calibration rows, PRINT comparison only (no writes)
  python3 rescore.py              # full rescore, writes to the tracker

Behavior:
  - LinkedIn rows: refetches the job description fresh (up to 6,000 chars) so
    fluff-truncated rows (like POOLCORP) are scored on the real JD. If the
    posting no longer exists, marks the Notes with a CLOSED? flag.
  - Adzuna/USAJOBS rows: rescored from the stored description (full-JD fetch
    for these arrives in M2.3).
  - Preserves ALL your edits: Status is never touched. For rows where Status
    is not "New" (you've acted on them), Notes and Resume Used are also
    preserved — only JD Match / Match % / Employer Type update.
"""

import re
import sys
import time
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

import agent  # reuses analyze_job_with_claude, match_tier, pick_resume, ensure_workbook

EXCEL_PATH = agent.EXCEL_PATH
SHEET      = agent.SHEET_NAME

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}


def check_closed_on_page(job_id):
    """
    Best-effort closed check against the public job page, where LinkedIn shows
    the 'No longer accepting applications' banner. Guest traffic sometimes hits
    a login wall — in that case we return False (unknown), never a false CLOSED.
    """
    try:
        resp = requests.get(f"https://www.linkedin.com/jobs/view/{job_id}/",
                            headers=HEADERS, timeout=15)
        if resp.status_code == 404:
            return True
        return "no longer accepting applications" in resp.text.lower()
    except Exception:
        return False


def refetch_linkedin(link):
    """Returns (description up to 6000 chars, closed_flag). None desc = keep stored."""
    m = re.search(r"/jobs/view/(\d+)", link or "")
    if not m:
        return None, False
    try:
        resp = requests.get(
            f"https://www.linkedin.com/jobs-guest/jobs/api/jobPosting/{m.group(1)}",
            headers=HEADERS, timeout=15)
        if resp.status_code == 404:
            return None, True
        if resp.status_code != 200:
            return None, False
        soup = BeautifulSoup(resp.text, "html.parser")
        # The API fragment doesn't carry the closed banner — check the public
        # job page for it (best-effort; login walls read as unknown/open).
        closed = check_closed_on_page(m.group(1))
        desc_el = soup.find("div", class_=lambda x: x and "description__text" in x)
        if not desc_el:
            return None, True  # page exists but no JD body — likely closed
        return desc_el.get_text(strip=True)[:6000], closed
    except Exception:
        return None, False


def main():
    sample_mode = "--sample" in sys.argv

    agent.ensure_workbook()  # runs the Employer Type migration if needed
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET]

    rows = []
    for r in range(2, ws.max_row + 1):
        if not ws.cell(row=r, column=2).value:
            continue
        rows.append({
            "r": r,
            "company":  ws.cell(row=r, column=2).value or "",
            "title":    ws.cell(row=r, column=3).value or "",
            "tier":     ws.cell(row=r, column=4).value or "",
            "pct":      ws.cell(row=r, column=5).value,
            "status":   ws.cell(row=r, column=7).value or "New",
            "loc":      ws.cell(row=r, column=9).value or "",
            "source":   ws.cell(row=r, column=11).value or "",
            "link":     ws.cell(row=r, column=12).value or "",
            "desc":     ws.cell(row=r, column=15).value or "",
        })

    if sample_mode:
        tops    = [x for x in rows if x["tier"] == "Top"][:3]
        highs   = [x for x in rows if x["tier"] == "High"]
        highs   = highs[:: max(1, len(highs) // 5)][:5]
        meds    = [x for x in rows if x["tier"] == "Medium"][:2]
        lows    = [x for x in rows if x["tier"] == "Low"][:1]
        pool    = [x for x in rows if "pool" in x["company"].lower()]
        targets = {x["r"]: x for x in tops + highs + meds + lows + pool}
        targets = list(targets.values())
        print(f"— SAMPLE MODE: {len(targets)} calibration rows, no writes —\n")
    else:
        targets = rows
        print(f"— FULL RESCORE: {len(targets)} rows —\n")

    changed = 0
    for i, x in enumerate(targets, start=1):
        desc, closed = x["desc"], False
        if x["source"] == "LinkedIn":
            fresh, closed = refetch_linkedin(x["link"])
            if fresh and len(fresh) > len(desc):
                desc = fresh

        job = {"title": x["title"], "company": {"display_name": x["company"]},
               "location": {"display_name": x["loc"]}, "description": desc}
        analysis = agent.analyze_job_with_claude(job)

        if analysis is None:
            print(f"[{i}/{len(targets)}] {x['title']} @ {x['company']} — scoring failed, skipped")
            continue

        comp = analysis.get("components") or {}
        try:
            pct = max(0, min(100, sum(int(comp.get(k, 0)) for k in
                     ("title_fit", "responsibilities", "requirements", "domain"))))
        except Exception:
            pct = analysis.get("match_percent", 0)
        dq = analysis.get("disqualifiers") or []
        if dq:
            pct = min(pct, 45)
        tier     = agent.match_tier(pct)
        employer = analysis.get("employer_type", "Unclear")
        if employer not in ("Direct", "Staffing", "Unclear"):
            employer = "Unclear"

        arrow = f"{x['pct']}% {x['tier']}  →  {pct}% {tier}"
        breakdown = f"[T{comp.get('title_fit','?')}/R{comp.get('responsibilities','?')}/Req{comp.get('requirements','?')}/D{comp.get('domain','?')}]"
        flag = "  ⚠️ POSTING MAY BE CLOSED" if closed else ""
        if dq:
            flag += f"  🚫 DQ: {'; '.join(str(d) for d in dq)}"
        print(f"[{i}/{len(targets)}] {x['title']} @ {x['company']} ({employer})")
        print(f"    {arrow}  {breakdown}{flag}")

        if not sample_mode:
            ws.cell(row=x["r"], column=4).value = tier
            ws.cell(row=x["r"], column=5).value = pct
            ws.cell(row=x["r"], column=16).value = employer
            if x["status"] == "New":  # untouched rows get full refresh
                ws.cell(row=x["r"], column=6).value = agent.pick_resume(analysis, x["title"])
                notes = analysis.get("notes", "") + f" {breakdown}"
                if dq:
                    notes = "DISQUALIFIER: " + "; ".join(str(d) for d in dq) + ". " + notes
                if closed:
                    notes = "⚠️ POSTING MAY BE CLOSED. " + notes
                ws.cell(row=x["r"], column=14).value = notes
                if len(desc) > len(x["desc"]):
                    ws.cell(row=x["r"], column=15).value = desc
            changed += 1
            if changed % 25 == 0:
                wb.save(EXCEL_PATH)
                print(f"    💾 Progress saved ({changed} rows)")

        time.sleep(1)

    if not sample_mode:
        wb.save(EXCEL_PATH)
        print(f"\n💾 Rescore complete — {changed} rows updated in {EXCEL_PATH}")
    else:
        print("\nSample complete — nothing written. Review the old→new movements above.")


if __name__ == "__main__":
    main()
