"""
Job Search Agent — v2 (M1)
Pulls jobs from Adzuna, LinkedIn, and USAJOBS; scores them with Claude Haiku;
appends new listings to Job_Tracker.xlsx in the repo; notifies via Slack.

Changes from v1:
  - Targets exactly 4 titles: PO, Sr PO, PM, Sr PM
  - USAJOBS added as a third source (official free API)
  - Notion removed — Excel file in the repo is the single source of truth
  - Dedup hardened: URL match + company|title|location fingerprint (cross-source)
  - New match tiers: Top 75-100, High 50-74, Medium 25-49, Low 0-24
  - Resume selection maps to the 3 real resume files (PM / PO / MES)
  - Scoring runs on Claude Haiku 4.5 (Groq free tier could not handle the
    token volume). Retries with backoff on rate limits; saves every 25 jobs.
"""

import os
import re
import json
import time
import requests
from datetime import datetime, timezone, timedelta
from dotenv import load_dotenv
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ── Load environment variables ────────────────────────────────────────────────
load_dotenv()

ADZUNA_APP_ID      = os.getenv("ADZUNA_APP_ID")
ADZUNA_APP_KEY     = os.getenv("ADZUNA_APP_KEY")
GROQ_API_KEY       = os.getenv("GROQ_API_KEY")        # reserved for M2 prefilter
ANTHROPIC_API_KEY  = os.getenv("ANTHROPIC_API_KEY")
USAJOBS_API_KEY    = os.getenv("USAJOBS_API_KEY")      # register free at developer.usajobs.gov
USAJOBS_USER_AGENT = os.getenv("USAJOBS_USER_AGENT")   # the email you registered with
SLACK_WEBHOOK_URL  = os.getenv("SLACK_WEBHOOK_URL")

EXCEL_PATH  = "Job_Tracker.xlsx"
SHEET_NAME  = "Tracker"
TRACKER_URL = "https://github.com/rontheaiguy/job-search-agent/blob/main/Job_Tracker.xlsx"

# ── Target job titles ─────────────────────────────────────────────────────────

JOB_TITLES = [
    "Product Owner",
    "Senior Product Owner",
    "Product Manager",
    "Senior Product Manager",
]

# Titles allowed through the filter (lowercase substring match)
ALLOWED_TITLES = [
    "product owner",
    "senior product owner",
    "sr. product owner",
    "sr product owner",
    "product manager",
    "senior product manager",
    "sr. product manager",
    "sr product manager",
]

# Title fragments that disqualify a listing even if an allowed phrase matches
BLOCKED_TITLE_FRAGMENTS = [
    "associate product manager",
    "staff product manager",
    "principal product manager",
    "group product manager",
    "director",
    "vp",
    "vice president",
    "head of product",
    "intern",
    "co-op",
]

# Resume files (must match the file names used when applying)
RESUME_PM  = "Rounaq_Gandhi_Resume_Product_Manager.pdf"
RESUME_PO  = "Rounaq_Gandhi_Resume_Product_Owner.pdf"
RESUME_MES = "Rounaq_Gandhi_Resume_MES.pdf"

# ── Excel column layout (must match Job_Tracker.xlsx headers) ─────────────────

COLUMNS = [
    "Date Added",            # A
    "Company",               # B
    "Title",                 # C
    "JD Match",              # D
    "Match %",               # E
    "Resume Used",           # F
    "Status",                # G
    "Onsite/Hybrid/Remote",  # H
    "Location",              # I
    "Salary",                # J
    "Source",                # K
    "JD Link",               # L
    "Key Skills",            # M
    "Notes",                 # N
    "Job Description",       # O
]

# ── Step 1: Fetch jobs ────────────────────────────────────────────────────────

def fetch_jobs_from_adzuna():
    """Calls the Adzuna Jobs API for each target title."""
    print("🔍 Fetching jobs from Adzuna...")
    all_jobs = []

    for title in JOB_TITLES:
        print(f"  → Searching: {title}")
        title_jobs = []

        for page in range(1, 6):  # 5 pages x 50 = up to 250 per title
            url = "https://api.adzuna.com/v1/api/jobs/us/search/" + str(page)
            params = {
                "app_id":           ADZUNA_APP_ID,
                "app_key":          ADZUNA_APP_KEY,
                "what":             title,
                "where":            "United States",
                "results_per_page": 50,
                "max_days_old":     2,   # Mon run covers the weekend gap
                "sort_by":          "date",
            }

            try:
                resp = requests.get(url, params=params, timeout=30)
                resp.raise_for_status()
                jobs = resp.json().get("results", [])
                if not jobs:
                    break
                title_jobs.extend(jobs)
                if len(jobs) < 50:
                    break
            except Exception as e:
                print(f"     ⚠️  Adzuna error on page {page} for '{title}': {e}")
                break

            time.sleep(0.5)

        print(f"     Found {len(title_jobs)} listings for '{title}'")
        all_jobs.extend(title_jobs)
        time.sleep(1)

    print(f"✅ Adzuna total: {len(all_jobs)}")
    return all_jobs


def fetch_jobs_from_linkedin():
    """Scrapes LinkedIn's public guest API for each target title."""
    print("🔍 Fetching jobs from LinkedIn...")
    all_jobs = []

    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }

    TIME_FILTER = "r172800"  # last 48 hours — Mon run covers the weekend gap

    for title in JOB_TITLES:
        print(f"  → Searching LinkedIn: {title}")
        title_jobs = []
        search_title = title.replace(" ", "%20")

        for page in range(5):
            start = page * 10
            url = (
                f"https://www.linkedin.com/jobs-guest/jobs/api/seeMoreJobPostings/search"
                f"?keywords={search_title}"
                f"&location=United%20States"
                f"&f_TPR={TIME_FILTER}"
                f"&start={start}"
            )

            try:
                resp = requests.get(url, headers=headers, timeout=15)
                if resp.status_code != 200:
                    break

                job_ids = re.findall(r'data-entity-urn="urn:li:jobPosting:(\d+)"', resp.text)
                if not job_ids:
                    break

                for job_id in job_ids:
                    detail_url = f"https://www.linkedin.com/jobs-guest/jobs/api/jobPosting/{job_id}"
                    try:
                        detail_resp = requests.get(detail_url, headers=headers, timeout=15)
                        if detail_resp.status_code != 200:
                            continue

                        soup = BeautifulSoup(detail_resp.text, "html.parser")

                        title_el    = soup.find("h2", class_=lambda x: x and "top-card-layout__title" in x)
                        company_el  = soup.find("a", class_=lambda x: x and "topcard__org-name-link" in x)
                        location_el = soup.find("span", class_=lambda x: x and "topcard__flavor--bullet" in x)
                        desc_el     = soup.find("div", class_=lambda x: x and "description__text" in x)

                        job_title   = title_el.get_text(strip=True) if title_el else ""
                        company     = company_el.get_text(strip=True) if company_el else ""
                        location    = location_el.get_text(strip=True) if location_el else ""
                        description = desc_el.get_text(strip=True)[:6000] if desc_el else ""

                        if not job_title:
                            continue

                        title_jobs.append({
                            "title":        job_title,
                            "company":      {"display_name": company},
                            "location":     {"display_name": location},
                            "description":  description,
                            "redirect_url": f"https://www.linkedin.com/jobs/view/{job_id}/",
                            "created":      datetime.now(timezone.utc).isoformat(),
                            "source":       "LinkedIn",
                        })

                        time.sleep(0.3)

                    except Exception:
                        continue

                time.sleep(1)

            except Exception as e:
                print(f"     ⚠️  LinkedIn error on page {page} for '{title}': {e}")
                break

        print(f"     Found {len(title_jobs)} listings for '{title}'")
        all_jobs.extend(title_jobs)
        time.sleep(2)

    print(f"✅ LinkedIn total: {len(all_jobs)}")
    return all_jobs


def fetch_jobs_from_usajobs():
    """
    Calls the official USAJOBS API for each target title.
    Free — register at developer.usajobs.gov for an API key.
    Skips gracefully if USAJOBS secrets are not configured yet.
    """
    if not USAJOBS_API_KEY or not USAJOBS_USER_AGENT:
        print("ℹ️  USAJOBS skipped — USAJOBS_API_KEY / USAJOBS_USER_AGENT not set.")
        return []

    print("🔍 Fetching jobs from USAJOBS...")
    all_jobs = []

    headers = {
        "Host":              "data.usajobs.gov",
        "User-Agent":        USAJOBS_USER_AGENT,
        "Authorization-Key": USAJOBS_API_KEY,
    }

    for title in JOB_TITLES:
        print(f"  → Searching USAJOBS: {title}")
        try:
            resp = requests.get(
                "https://data.usajobs.gov/api/search",
                headers=headers,
                params={
                    "PositionTitle":  title,
                    "ResultsPerPage": 100,
                    "SortField":      "OpenDate",
                    "SortDirection":  "Desc",
                },
                timeout=30,
            )
            resp.raise_for_status()
            items = resp.json().get("SearchResult", {}).get("SearchResultItems", [])

            count = 0
            for item in items:
                d = item.get("MatchedObjectDescriptor", {})

                # Salary from PositionRemuneration
                salary = ""
                rem = d.get("PositionRemuneration", [])
                if rem:
                    lo = rem[0].get("MinimumRange", "")
                    hi = rem[0].get("MaximumRange", "")
                    interval = rem[0].get("RateIntervalCode", "")
                    if lo and hi:
                        salary = f"${float(lo):,.0f} – ${float(hi):,.0f} ({interval})"

                summary = d.get("UserArea", {}).get("Details", {}).get("JobSummary", "")

                all_jobs.append({
                    "title":        d.get("PositionTitle", ""),
                    "company":      {"display_name": d.get("OrganizationName", "US Government")},
                    "location":     {"display_name": d.get("PositionLocationDisplay", "")},
                    "description":  summary[:6000],
                    "redirect_url": d.get("PositionURI", ""),
                    "created":      d.get("PublicationStartDate", ""),
                    "salary":       salary,
                    "source":       "USAJOBS",
                })
                count += 1

            print(f"     Found {count} listings for '{title}'")
            time.sleep(1)

        except Exception as e:
            print(f"     ⚠️  USAJOBS error for '{title}': {e}")

    print(f"✅ USAJOBS total: {len(all_jobs)}")
    return all_jobs

# ── Step 2: Filter ────────────────────────────────────────────────────────────

def is_recent(job):
    """Keeps only listings posted in the last 3 days (safety net)."""
    raw_date = job.get("created") or ""
    if not raw_date:
        return True

    try:
        if "T" in raw_date:
            posted = datetime.fromisoformat(raw_date.replace("Z", "+00:00"))
        else:
            posted = datetime.strptime(raw_date[:10], "%Y-%m-%d").replace(tzinfo=timezone.utc)
        return posted >= datetime.now(timezone.utc) - timedelta(days=3)
    except Exception:
        return True


def is_relevant_title(job):
    """Allows only the 4 target titles (plus Sr./Sr spelling variants)."""
    title = job.get("title", "").lower().strip()

    for blocked in BLOCKED_TITLE_FRAGMENTS:
        if blocked in title:
            return False

    return any(allowed in title for allowed in ALLOWED_TITLES)


def normalize(text):
    """Lowercase, strip punctuation and extra spaces — used for fingerprints."""
    return re.sub(r"[^a-z0-9 ]", "", (text or "").lower()).strip()


def fingerprint(company, title, location):
    """
    Cross-source duplicate KEY: company|title|location.
    A fingerprint match alone is NOT enough to call something a duplicate —
    the descriptions must also substantially overlap (see desc_similar).
    This keeps multiple genuine openings at the same company/title/location
    (distinct reqs are written differently) while collapsing the same job
    cross-posted on two boards (same text).
    """
    return f"{normalize(company)}|{normalize(title)}|{normalize(location)}"


def desc_tokens(text):
    """Word-token set from the first 800 chars of a description."""
    return set(re.findall(r"[a-z0-9]+", (text or "")[:800].lower()))


def extract_req_id(text):
    """
    Pulls the employer's requisition ID from description text when present
    (e.g. 'R0000378340', '730485BR', 'Req ID: 12345', 'Job ID # JR-98765').
    Board-independent, so it's decisive across sources: same req ID = same
    job; different req IDs = definitely separate openings.
    Returns '' if no ID found.
    """
    if not text:
        return ""
    # Labeled patterns: "Req ID: X", "Requisition Number X", "Job ID # X"
    m = re.search(
        r"(?:req(?:uisition)?|job|position)\s*(?:id|number|no\.?|#)?\s*[:#]?\s*([A-Z]{0,3}[-_]?\d{4,}[A-Z]{0,3})",
        text, re.IGNORECASE)
    if m:
        return m.group(1).upper()
    # Bare Workday-style (R0000378340) or BR-style (730485BR) IDs
    m = re.search(r"\b(R\d{7,}|\d{5,}BR)\b", text)
    return m.group(1).upper() if m else ""


def desc_similar(tokens_a, tokens_b):
    """
    True if two descriptions substantially overlap (containment ratio >= 0.65).
    Containment (not symmetric overlap) is used because Adzuna often carries a
    truncated snippet of the same posting LinkedIn shows in full.
    If either description is missing, we can't distinguish → treat as similar
    (prefer skipping a possible dupe over flooding the tracker).
    """
    if not tokens_a or not tokens_b:
        return True
    inter = len(tokens_a & tokens_b)
    return inter / min(len(tokens_a), len(tokens_b)) >= 0.65


def read_existing_keys():
    """
    Reads Job_Tracker.xlsx and returns:
      - set of JD links
      - dict of fingerprint -> list of description token sets
    so we never re-add a job that's already in the tracker.
    """
    print("📋 Reading existing tracker entries...")
    links, prints = set(), {}

    if not os.path.exists(EXCEL_PATH):
        print("  ⚠️  Tracker file not found — starting with empty dedup sets.")
        return links, prints

    wb = load_workbook(EXCEL_PATH, read_only=True)
    ws = wb[SHEET_NAME]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        company  = row[1] or ""    # B
        title    = row[2] or ""    # C
        location = row[8] or ""    # I
        link     = (row[11] or "").strip()  # L
        desc     = row[14] or ""   # O
        if link:
            links.add(link)
        fp = fingerprint(company, title, location)
        prints.setdefault(fp, []).append(
            {"tokens": desc_tokens(desc), "req_id": extract_req_id(desc)})

    wb.close()
    print(f"  → {len(links)} existing listings in tracker.")
    return links, prints


def is_duplicate(fp, tokens, req_id, prints):
    """
    Duplicate decision for a fingerprint (company|title|location) match:
      1. Both listings have a req ID → the IDs decide (same = dupe, different = keep)
      2. Otherwise → description similarity decides
    """
    for stored in prints.get(fp, []):
        if req_id and stored["req_id"]:
            if req_id == stored["req_id"]:
                return True
            continue  # different req IDs = definitely separate openings
        if desc_similar(tokens, stored["tokens"]):
            return True
    return False


def filter_jobs(raw_jobs, existing_links, existing_prints):
    """
    Removes irrelevant titles, stale posts, and duplicates.
    Duplicate = same URL, OR same company|title|location fingerprint AND
    substantially similar description. Same fingerprint with a clearly
    different description = a separate opening → kept.
    """
    print("🔎 Filtering listings...")
    filtered = []
    seen_links = set()
    seen_prints = {fp: list(v) for fp, v in existing_prints.items()}
    skipped_title = skipped_dupe = 0

    for job in raw_jobs:
        link = (job.get("redirect_url") or "").strip()
        if not link:
            continue

        if not is_relevant_title(job):
            skipped_title += 1
            continue

        if not is_recent(job):
            continue

        company  = job.get("company", {}).get("display_name") or ""
        title    = job.get("title", "")
        location = job.get("location", {}).get("display_name") or ""
        fp       = fingerprint(company, title, location)
        tokens   = desc_tokens(job.get("description"))
        req_id   = extract_req_id(job.get("description"))

        if (link in seen_links or link in existing_links
                or is_duplicate(fp, tokens, req_id, seen_prints)):
            skipped_dupe += 1
            continue

        seen_links.add(link)
        seen_prints.setdefault(fp, []).append({"tokens": tokens, "req_id": req_id})
        filtered.append(job)

    print(f"  → Skipped {skipped_title} irrelevant titles, {skipped_dupe} duplicates.")
    print(f"✅ {len(filtered)} new listings after filtering.")
    return filtered

# ── Step 3: Scoring (Groq/Llama — free tier; M2 upgrades this to hybrid) ─────

RESUME_PM_SUMMARY = """
RESUME A — Rounaq_Gandhi_Resume_Product_Manager.pdf (for PM / Sr PM roles)
Product Manager, 7+ years product management & ownership on a decade of software delivery.
PEEK (B2B SaaS, iOS) PM 04/2023-06/2026: owned strategy/GTM/launch of mobile POS barcode
scanner — lead differentiator in $18M GMV enterprise deals; shipped Offline Mode 2 weeks early
(22% adoption lift, closed top churn driver); retained $1.1M annual revenue via QR check-in and
kiosk redesign; 11.5% user adoption lift from discovery-driven enhancements; A/B testing with
PostHog/Looker (~$125K GR impact); PRDs, user stories, BDD/Gherkin, OKRs/KPIs, sprint
ceremonies; cut PRD drafting ~25% with AI workflows. Prior: Senior QE at Peek; Emerson
(pharma MES, Associate PO, $4.2M module); Cognizant. Certs: CSPO, CSM, SAFe Agilist.
""".strip()

RESUME_PO_SUMMARY = """
RESUME B — Rounaq_Gandhi_Resume_Product_Owner.pdf (for PO / Sr PO roles)
Product Owner / Product Manager, 7+ years. Same PEEK experience as Resume A but framed
around backlog ownership: owned the product backlog and full delivery lifecycle — user stories,
BDD/Gherkin acceptance criteria, PRDs, backlog prioritization from client interviews/NPS/support
escalations, sprint planning, refinement, retrospectives; owned the Offline Mode backlog decision
over two competing enterprise requests. Emerson: Associate Product Owner on Syncade MES —
requirements, traceability, release sign-off across full SDLC. Certs: CSPO, CSM, SAFe Agilist.
""".strip()

RESUME_MES_SUMMARY = """
RESUME C — Rounaq_Gandhi_Resume_MES.pdf (for MES / pharma / regulated-industry companies)
Product Owner/PM with 4+ years on Syncade MES in pharmaceutical manufacturing: authored and
configured master recipes executing as Electronic Batch Records (EBR) in GMP environment; owned
requirements, traceability (RTM), release sign-off under 21 CFR Part 11, IEC 62304, ISO 13485;
lead tester for Review-by-Exception (QRM) module — 200+ defects pre-release, $4.2M revenue,
customer-satisfaction award; ALCOA+. Plus recent B2B SaaS PM experience at Peek ($18M GMV,
$1.1M retained). Use whenever the company is pharma, med-device, MES, or FDA-regulated.
""".strip()


SCORING_SYSTEM = f"""You are a job-search assistant for Rounaq Gandhi. You analyze job listings against his three resume variants.

{RESUME_PM_SUMMARY}

{RESUME_PO_SUMMARY}

{RESUME_MES_SUMMARY}

For every job listing you receive, respond ONLY with a valid JSON object, no markdown, no extra text:
{{
  "industry": "<e.g. B2B SaaS, FinTech, Pharma, Enterprise Software>",
  "match_percent": <0-100 integer — how well the BEST resume fits this JD>,
  "key_skills": ["<skill1>", "<skill2>", "<skill3>"],
  "resume_file": "<exactly one of: {RESUME_PM}, {RESUME_PO}, {RESUME_MES}>",
  "notes": "<2-3 sentences: fit assessment, red flags, what to customize>"
}}

Resume selection rules (apply in this order):
1. If the company/domain is pharma, medical devices, MES, or FDA-regulated manufacturing → {RESUME_MES}
2. Else if the job title contains "Product Owner" → {RESUME_PO}
3. Else → {RESUME_PM}

Scoring guidance:
- Be conservative. Only score 75+ when the JD's core requirements clearly map to resume experience.
- Deduct for: hard requirements he lacks (e.g. specific industry depth, 8+ years in PM title, people management), clearance he'd need, niche technical stacks.
- match_percent bands: 75-100 Top, 50-74 High, 25-49 Medium, 0-24 Low"""


def analyze_job_with_claude(job):
    """
    Scores the job with Claude Haiku 4.5. The static context (resumes + rules)
    is sent as a cached system prompt, so per-job cost is only the JD itself.
    Retries with exponential backoff on rate limits / server errors.
    Returns None if scoring ultimately fails (caller records it as unscored).
    """
    title       = job.get("title", "Unknown Title")
    company     = job.get("company", {}).get("display_name") or "Unknown Company"
    description = (job.get("description") or "")[:5000]
    location    = job.get("location", {}).get("display_name") or ""

    user_msg = (f"Job listing:\nTitle: {title}\nCompany: {company}\n"
                f"Location: {location}\nDescription: {description}")

    for attempt in range(4):
        try:
            response = requests.post(
                "https://api.anthropic.com/v1/messages",
                headers={
                    "x-api-key":         ANTHROPIC_API_KEY,
                    "anthropic-version": "2023-06-01",
                    "content-type":      "application/json",
                },
                json={
                    "model":       "claude-haiku-4-5",
                    "max_tokens":  600,
                    "temperature": 0,  # deterministic — same job always scores the same
                    "system": [{
                        "type": "text",
                        "text": SCORING_SYSTEM,
                        "cache_control": {"type": "ephemeral"},
                    }],
                    "messages": [{"role": "user", "content": user_msg}],
                },
                timeout=60,
            )

            if response.status_code in (429, 500, 502, 503, 529):
                wait = int(response.headers.get("retry-after", 0)) or (5 * (2 ** attempt))
                print(f"     ⏳ API busy ({response.status_code}) — retrying in {wait}s...")
                time.sleep(wait)
                continue

            response.raise_for_status()
            raw_text = response.json()["content"][0]["text"].strip()

            if raw_text.startswith("```"):
                raw_text = raw_text.split("```")[1]
                if raw_text.startswith("json"):
                    raw_text = raw_text[4:]
            return json.loads(raw_text.strip())

        except Exception as e:
            print(f"     ⚠️  Claude API error for '{title}' (attempt {attempt + 1}): {e}")
            time.sleep(5 * (2 ** attempt))

    return None  # scoring failed after all retries


def match_tier(percent):
    """Tier thresholds: Top 75-100, High 50-74, Medium 25-49, Low 0-24."""
    if percent >= 75:
        return "Top"
    elif percent >= 50:
        return "High"
    elif percent >= 25:
        return "Medium"
    else:
        return "Low"


def pick_resume(analysis, job_title):
    """Validates the LLM's resume pick; falls back to deterministic title rule."""
    pick = (analysis.get("resume_file") or "").strip()
    if pick in (RESUME_PM, RESUME_PO, RESUME_MES):
        return pick
    return RESUME_PO if "owner" in job_title.lower() else RESUME_PM


def work_mode(job):
    """Guesses Remote / Hybrid / Onsite from location string and description."""
    text = ((job.get("location", {}).get("display_name") or "") + " "
            + (job.get("description") or "")[:300]).lower()
    if "remote" in text:
        return "Remote"
    elif "hybrid" in text:
        return "Hybrid"
    return "Onsite"


def extract_salary(job):
    """Salary from USAJOBS field or Adzuna salary_min/max; blank otherwise."""
    if job.get("salary"):
        return job["salary"]
    lo, hi = job.get("salary_min"), job.get("salary_max")
    if lo and hi:
        return f"${lo:,.0f} – ${hi:,.0f}"
    if lo:
        return f"${lo:,.0f}+"
    return ""


def detect_source(job):
    """Source from explicit tag or the URL domain."""
    if job.get("source"):
        return job["source"]
    link = (job.get("redirect_url") or "").lower()
    for key, name in [
        ("linkedin.com", "LinkedIn"), ("usajobs.gov", "USAJOBS"),
        ("indeed.com", "Indeed"), ("greenhouse.io", "Greenhouse"),
        ("lever.co", "Lever"), ("workday", "Workday"),
        ("ziprecruiter.com", "ZipRecruiter"), ("glassdoor.com", "Glassdoor"),
        ("adzuna.com", "Adzuna"), ("smartrecruiters.com", "SmartRecruiters"),
        ("icims.com", "iCIMS"),
    ]:
        if key in link:
            return name
    return "Adzuna"  # Adzuna aggregates — its redirect links vary

# ── Step 4: Append to Excel ───────────────────────────────────────────────────

def ensure_workbook():
    """Creates Job_Tracker.xlsx with headers + dropdowns if it doesn't exist."""
    if os.path.exists(EXCEL_PATH):
        return

    print(f"  ℹ️  {EXCEL_PATH} not found — creating it.")
    widths = [12, 24, 32, 10, 9, 38, 13, 20, 24, 22, 14, 45, 30, 45, 60]
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    font = Font(color="FFFFFF", bold=True, size=11)
    for i, (col, w) in enumerate(zip(COLUMNS, widths), start=1):
        c = ws.cell(row=1, column=i, value=col)
        c.fill, c.font = fill, font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    ws.row_dimensions[1].height = 28

    validations = [
        ("D", '"Top,High,Medium,Low"'),
        ("F", f'"{RESUME_PM},{RESUME_PO},{RESUME_MES}"'),
        ("G", '"New,Applied,Interviewing,Rejected,Offer,Skipped"'),
        ("H", '"Onsite,Hybrid,Remote"'),
    ]
    for col, formula in validations:
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}5000")

    wb.save(EXCEL_PATH)


def append_to_excel(rows):
    """
    Appends new job rows to Job_Tracker.xlsx (creating it if missing).
    NEVER modifies existing rows — your Status/Notes edits are preserved.
    """
    if not rows:
        return 0

    ensure_workbook()
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    for r in rows:
        ws.append([
            r["date_added"], r["company"], r["title"], r["tier"], r["match_percent"],
            r["resume"], "New", r["mode"], r["location"], r["salary"],
            r["source"], r["link"], r["key_skills"], r["notes"], r["description"],
        ])

    wb.save(EXCEL_PATH)
    return len(rows)

# ── Step 5: Slack summary ─────────────────────────────────────────────────────

def send_slack_summary(saved_jobs):
    """Posts the morning digest to Slack."""
    today = datetime.now().strftime("%m/%d/%Y")

    if not saved_jobs:
        message = f"📋 *Job Search — {today}*\n\nNo new listings found today."
    else:
        total = len(saved_jobs)
        tier_counts = {"Top": 0, "High": 0, "Medium": 0, "Low": 0}
        for j in saved_jobs:
            tier_counts[j["tier"]] = tier_counts.get(j["tier"], 0) + 1

        top_matches = sorted(
            [j for j in saved_jobs if j["tier"] in ("Top", "High")],
            key=lambda x: x["match_percent"], reverse=True,
        )[:5]

        lines = [
            f"📋 *Job Search — {today}*\n",
            f"✅ *{total} new listing{'s' if total != 1 else ''} added to the tracker*\n",
        ]
        if top_matches:
            lines.append("🔥 *Top matches:*")
            for j in top_matches:
                lines.append(
                    f"• <{j['link']}|{j['title']}> at {j['company']} — "
                    f"{j['match_percent']}% ({j['tier']}) → {j['resume'].replace('Rounaq_Gandhi_Resume_', '').replace('.pdf', '')}"
                )
            lines.append("")
        lines.append("📊 *Breakdown:*")
        lines.append(f"• 🏆 Top (75–100): {tier_counts['Top']}")
        lines.append(f"• 🟢 High (50–74): {tier_counts['High']}")
        lines.append(f"• 🟡 Medium (25–49): {tier_counts['Medium']}")
        lines.append(f"• 🔴 Low (0–24): {tier_counts['Low']}")
        lines.append(f"\n🔗 <{TRACKER_URL}|Open Job Tracker (Excel)>")
        message = "\n".join(lines)

    try:
        resp = requests.post(SLACK_WEBHOOK_URL, json={"text": message}, timeout=10)
        print("✅ Slack notification sent." if resp.status_code == 200
              else f"⚠️  Slack error: {resp.status_code} — {resp.text}")
    except Exception as e:
        print(f"⚠️  Slack error: {e}")

# ── Main orchestrator ─────────────────────────────────────────────────────────

def main():
    print("\n========================================")
    print("  Job Search Agent v2 — Starting Run")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("========================================\n")

    # 1. Pull raw jobs from all three sources
    raw_jobs = (fetch_jobs_from_adzuna()
                + fetch_jobs_from_linkedin()
                + fetch_jobs_from_usajobs())
    print(f"\n📦 Combined total: {len(raw_jobs)} raw listings")

    # 2. Read existing tracker entries for dedup
    existing_links, existing_prints = read_existing_keys()

    # 3. Filter
    new_jobs = filter_jobs(raw_jobs, existing_links, existing_prints)

    if not new_jobs:
        print("\nℹ️  No new listings to process.")
        send_slack_summary([])
        return

    # 4. Score each job and build rows (saved to Excel every 25 jobs so a
    #    crash or interruption never loses more than the current batch)
    all_rows, buffer, failed = [], [], 0
    for i, job in enumerate(new_jobs, start=1):
        title   = job.get("title", "Unknown Title")
        company = job.get("company", {}).get("display_name") or "Unknown Company"
        print(f"\n[{i}/{len(new_jobs)}] Analyzing: {title} at {company}")

        analysis = analyze_job_with_claude(job)

        if analysis is None:
            failed += 1
            pct, tier = "", ""
            resume = pick_resume({}, title)
            key_skills, notes = "", "Scoring failed — run again later or score manually."
            print("     ❌ Scoring failed — recorded unscored.")
        else:
            pct    = analysis.get("match_percent", 0)
            tier   = match_tier(pct)
            resume = pick_resume(analysis, title)
            key_skills = ", ".join(analysis.get("key_skills", []))
            notes  = analysis.get("notes", "")
            print(f"     Match: {pct}% ({tier}) | Resume: {resume}")

        buffer.append({
            "date_added":    datetime.now().strftime("%Y-%m-%d"),
            "company":       company,
            "title":         title,
            "tier":          tier,
            "match_percent": pct,
            "resume":        resume,
            "mode":          work_mode(job),
            "location":      job.get("location", {}).get("display_name") or "",
            "salary":        extract_salary(job),
            "source":        detect_source(job),
            "link":          (job.get("redirect_url") or "").strip(),
            "key_skills":    key_skills,
            "notes":         notes,
            "description":   (job.get("description") or "")[:5000],
        })

        if len(buffer) >= 25:
            append_to_excel(buffer)
            all_rows.extend(buffer)
            buffer = []
            print(f"     💾 Progress saved ({len(all_rows)} rows so far)")

        time.sleep(1)

    # 5. Flush remaining rows
    if buffer:
        append_to_excel(buffer)
        all_rows.extend(buffer)

    print(f"\n💾 {len(all_rows)} rows appended to {EXCEL_PATH}"
          + (f" ({failed} unscored — will not be re-fetched; score manually or ask Claude)" if failed else ""))

    # 6. Slack digest (unscored rows excluded from tier counts)
    send_slack_summary([r for r in all_rows if r["tier"]])

    print("\n========================================")
    print(f"  Run complete. {len(all_rows)} new listings saved.")
    print("========================================\n")


if __name__ == "__main__":
    main()
