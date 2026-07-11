"""
Merge Manual_JobTracker.xlsx into Job_Tracker.xlsx — safely.

Usage (both files in the project folder):
  python3 merge_manual.py

What it does:
  1. BACKS UP Job_Tracker.xlsx to Job_Tracker_backup_<timestamp>.xlsx first.
  2. Matches manual rows to tracker rows by LinkedIn job ID / URL, then by
     normalized company+title.
  3. Matched rows: fills in Status from the manual file ONLY where the tracker
     still says "New" (your tracker edits always win); appends manual notes.
  4. Unmatched rows: appended as new rows with Source = "Manual", unscored.
     If the JD-link cell contains pasted text instead of a URL, the text goes
     into Job Description where it belongs.
  5. Prints a full report. Nothing is silent.

Status mapping for manual rows: blank -> Applied, "Active" -> Interviewing,
"No" -> Rejected, anything else kept as-is.
"""

import os
import re
import shutil
from datetime import datetime
from openpyxl import load_workbook

MAIN, MANUAL = "Job_Tracker.xlsx", "Manual_JobTracker.xlsx"
SHEET = "Tracker"


def norm(s):
    return re.sub(r"[^a-z0-9 ]", "", (s or "").lower()).strip()


def li_id(url):
    m = re.search(r"/jobs/view/(\d+)", url or "")
    return m.group(1) if m else None


def map_status(s):
    s = (s or "").strip()
    if not s:
        return "Applied"
    low = s.lower()
    if low == "active":
        return "Interviewing"
    if low == "no":
        return "Rejected"
    return s


def main():
    for f in (MAIN, MANUAL):
        if not os.path.exists(f):
            print(f"❌ {f} not found in this folder — aborting, nothing touched.")
            return

    backup = f"Job_Tracker_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    shutil.copy2(MAIN, backup)
    print(f"🛟 Backup written: {backup}\n")

    wb = load_workbook(MAIN)
    ws = wb[SHEET]

    # Index existing tracker rows
    by_li, by_fp = {}, {}
    for r in range(2, ws.max_row + 1):
        co = ws.cell(row=r, column=2).value
        if not co:
            continue
        url = ws.cell(row=r, column=12).value or ""
        jid = li_id(url)
        if jid:
            by_li.setdefault(jid, r)
        fp = norm(co) + "|" + norm(ws.cell(row=r, column=3).value)
        by_fp.setdefault(fp, []).append(r)

    mwb = load_workbook(MANUAL)
    mws = mwb.worksheets[0]

    updated, appended, kept, ambiguous = 0, 0, 0, []

    for mr in mws.iter_rows(min_row=2, values_only=True):
        if not mr or not mr[0]:
            continue
        company, title = mr[0] or "", mr[1] or ""
        link_cell = (mr[2] or "").strip() if isinstance(mr[2], str) else ""
        status = map_status(mr[3] if len(mr) > 3 else "")
        mode = (mr[4] or "") if len(mr) > 4 else ""
        loc = (mr[5] or "") if len(mr) > 5 else ""
        notes = (mr[6] or "") if len(mr) > 6 else ""
        salary = (mr[7] or "") if len(mr) > 7 else ""

        is_url = link_cell.lower().startswith("http")
        pasted_jd = "" if is_url else link_cell[:5000]
        url = link_cell if is_url else ""

        # Find a match
        row = None
        jid = li_id(url)
        if jid and jid in by_li:
            row = by_li[jid]
        else:
            fp = norm(company) + "|" + norm(title)
            hits = by_fp.get(fp, [])
            if len(hits) == 1:
                row = hits[0]
            elif len(hits) > 1:
                ambiguous.append(f"{company} — {title} (matches {len(hits)} tracker rows; updated first)")
                row = hits[0]

        if row:
            cur = ws.cell(row=row, column=7).value or "New"
            if cur == "New":
                ws.cell(row=row, column=7).value = status
                updated += 1
            else:
                kept += 1
            if notes:
                old = ws.cell(row=row, column=14).value or ""
                if notes not in old:
                    ws.cell(row=row, column=14).value = (old + " | MANUAL: " + notes).strip(" |")
        else:
            ws.append([
                datetime.now().strftime("%Y-%m-%d"), company, title, "", "",
                "", status, mode, loc, salary, "Manual", url, "",
                ("MANUAL IMPORT. " + notes).strip(), pasted_jd, "",
            ])
            appended += 1

    wb.save(MAIN)

    print(f"✅ Merge complete:")
    print(f"   {updated} tracker rows updated with your manual status")
    print(f"   {kept} matches skipped (tracker already had a non-New status — your edits win)")
    print(f"   {appended} manual-only jobs appended (Source = Manual, unscored)")
    if ambiguous:
        print(f"\n⚠️  {len(ambiguous)} ambiguous matches (same company+title more than once):")
        for a in ambiguous:
            print("   -", a)
    print(f"\nIf anything looks wrong, restore with:  cp {backup} {MAIN}")


if __name__ == "__main__":
    main()
